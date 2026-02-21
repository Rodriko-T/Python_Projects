#Rodrick Tata Projects
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import os

def process_wb(filename, discount_percent):
    # If user didn't add .xlsx, add it automatically
    if not filename.lower().endswith(".xlsx"):
        filename = filename + ".xlsx"

    # Check if file exists
    if not os.path.exists(filename):
        print(f"{filename} not found in this directory")
        return

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Find the last column and add a new one
    last_col = sheet.max_column
    new_col = last_col + 1

    sheet.cell(row=1, column=new_col).value = 'New Prices'

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # assuming original price is in column 3
        corrected_price = cell.value * (1 - discount_percent / 100)

        corrected_price_cell = sheet.cell(row, new_col)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=new_col,
              max_col=new_col)

    Chart = BarChart()
    Chart.add_data(values)
    sheet.add_chart(Chart, sheet.cell(row=2, column=new_col + 1).coordinate)

    wb.save(filename)
    print(f"Your file, {filename} has been succesfully processed and saved in this directory")


# ===== USER INPUT PART =====

files_input = input("Enter Excel file names separated by commas: ")
discount_percent = float(input("Enter discount percentage (e.g., 10 for 10%): "))

files = [f.strip() for f in files_input.split(",")]

for file in files:
    process_wb(file, discount_percent)